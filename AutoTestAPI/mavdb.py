# Copyright (C) RflySim from Central South University and
# rflylab from School of Automation Science and Electrical Engineering, Beihang University.
# All Rights Reserved.
from unittest import result
import PX4MavCtrlV4 as PX4MavCtrl
import numpy as np
import re
import sqlite3
import os
import sys
import json
from fnmatch import fnmatch
import shutil
import pandas as pd
from openpyxl import Workbook

def dict_factory(cursor, row):  
    d = {}  
    for idx, col in enumerate(cursor.description):  
        d[col[0]] = row[idx]  
    return d 

class mavdb:
    def __init__(self):
        self.cursor = 0
        self.mydb = 0
        self.is_tested = 0
        self.dbpath = os.getcwd()
        self.jsonpath = sys.path[0]
        mavdb.json_loadto_sql(self)
        self.isvision = mavdb.vision(self)
        
    def vision(self):
        jsonpath = self.jsonpath  + '/db.json'
        with open(jsonpath, "r",encoding='utf-8') as f:
            db_data = json.load(f)
        isVision = db_data.get('Vision') # 为字典，需将字典的键提取出来形成字符串列表
        if isVision == 'on':
            return 1
        else:
            return 0


    # 获取游标
    def get_cursor(self):
        self.mydb = sqlite3.connect("{}//fault.db".format(self.dbpath))
        self.mydb.row_factory = dict_factory
        self.cursor=self.mydb.cursor()

     # 获取故障测试用例
    def get_fault_case(self):
        # 获取游标
        mavdb.get_cursor(self)
        sql='''
        select   *
        from     faultcase
        '''
        self.cursor.execute(sql)
        result=self.cursor.fetchall()
        return result

    # 获取故障测试用例ID列表
    def get_fault_case_ID(self):
        result = mavdb.get_fault_case(self)
        # 获取用户测试用例ID列表caselist
        path = sys.path[0]+'/db.json'

        with open(path, "r", encoding='utf-8') as f:
            db_data = json.load(f)

        caselist = []
        unique_index = []
        if db_data.get('testcase') == 'all':
            # 提取faultcase的CaseID
            casedata = result
            for data in casedata:
                ID = data.get('CaseID')
                caselist.append(ID)
                unique_index.append(data.get('UniqueIndex'))
        else:
            caselist = re.split(',', db_data.get('testcase'))  # ['3', '1', '30']
            caselist = [int(val) for val in caselist]  # [3, 1, 30]
            print(caselist)
            for i in range(len(caselist)):
                unique_index.append(db_data['faultcase'][caselist[i]-1]['UniqueIndex'])
        
        return caselist, unique_index

    # 处理控制序列
    def ctrl_seq_pro(self,case_id):
        # 获取游标
        mavdb.get_cursor(self)
        sql = '''
        select * from faultcase
        where CaseID = {}
        '''.format(case_id)
        self.cursor.execute(sql)
        data = self.cursor.fetchall()
        case_sequence = data[0].get('ControlSequence')
        case = re.split(';',case_sequence)
        cmd = np.array([])
        for i in range(len(case)):
            cmd = np.append(cmd,case[i])
        return cmd


    # 处理测试结果库（添加测试结果）
    def test_result_pro(self,data):
        # 1、获取游标
        mavdb.get_cursor(self)
        # 2、插入测试结果
        sql = '''
        insert into testresult
        (CaseID,FaultTestType,FaultParameters,TestStatus,SafetyAssessmentResults)
        values(?,?,?,?,?) 
        '''
        value = (data[0],data[1],data[2],data[3],data[4])
        # 第三步，执行
        self.cursor.execute(sql,value)
        # 第四步，提交
        self.mydb.commit()
    
    # 处理测试用例库（更改测试状态）
    def fault_case_pro(self,case_id):
        # 1、获取游标
        mavdb.get_cursor(self)
        # 2、更改测试状态
        sql = '''
        update faultcase
        set TestStatus = 'Finished'
        where CaseID = {}
        '''.format(case_id)
        # 第三步，执行
        self.cursor.execute(sql)
        # 第四步，提交
        self.mydb.commit()

    # 判断是否插入了重复的CaseID
    def is_TheSame_ID_pro(self,case_id):
        # 1、获取游标
        mavdb.get_cursor(self)
        sql = '''
        select * from faultcase
        where CaseID = {}
        '''.format(case_id)
        self.cursor.execute(sql)
        data = self.cursor.fetchall()
        # 如果插入同一个ID的用例，则删除用例库中的用例，插入新的
        if len(data) != 0:
            sql2 = '''
            delete from faultcase
            where CaseID = {}
            '''.format(case_id)
            self.cursor.execute(sql2)
            self.mydb.commit()

    #判断是否为已测试过用例 
    def is_tested_pro(self,case_id):
        # 1、获取游标
        mavdb.get_cursor(self)
        sql = '''
        select * from faultcase
        where CaseID = {}
        '''.format(case_id)
        self.cursor.execute(sql)
        data = self.cursor.fetchall()

        # 如果已测试过，则删除结果库的记录
        if data[0].get('TestStatus') == 'Finished':
            self.is_tested = 1
        else:
            self.is_tested = 0
        return self.is_tested

    # 处理重复用例测试的结果库
    def repeat_case_pro(self,case_id):
        # 1、获取游标
        mavdb.get_cursor(self)
        sql = '''
        select * from testresult
        where CaseID = {}
        '''.format(case_id)
        self.cursor.execute(sql)
        data = self.cursor.fetchall()
        print(len(data))

        # 如果已测试过，则删除结果库的记录
        if len(data) != 0:
            sql2 = '''
            delete from testresult
            where CaseID = {}
            '''.format(case_id)
            self.cursor.execute(sql2)
            self.mydb.commit()

    def json_loadto_sql(self):
        path = self.jsonpath + '/db.json'
        db_case = mavdb.get_fault_case(self)

        with open(path, "r",encoding='utf-8') as f:
            db_data = json.load(f)

        json_case = db_data.get('faultcase')

        # 找出相同CaseID的用例并比较，覆盖相同ID的不同用例,同时补充json中剩余的用例到用例库
        # 1、提取json的CaseID
        if len(json_case) >= 1:
            self.json_ID = []
            for i in range(len(json_case)):
                # print(json_case[i])
                self.json_ID.append(json_case[i].get("CaseID"))

            # 2、提取数据库中的CaseID
            self.db_ID = []
            for i in range(len(db_case)):
                # print(db_case[i])
                self.db_ID.append(db_case[i].get("CaseID"))

            # 3、取交集并转化为list
            self.con_ID = list(set(self.json_ID)&set(self.db_ID))

            # 4、取出两者交集ID
            con_json_case = []
            for i in range(len(json_case)):
                if json_case[i].get('CaseID') in self.con_ID:
                    con_json_case.append(json_case[i])

            con_db_case = []
            for i in range(len(db_case)):
                if db_case[i].get('CaseID') in self.con_ID:
                    con_db_case.append(db_case[i])

            # 5、比较两者的case是否相同，如果相同，则覆盖掉用例库中的用例
            for i in range(len(con_json_case)):
                # 如果json的数据不等于数据库中的，则覆盖掉数据库中的
                if con_db_case[i] != con_json_case[i]:
                    # print(con_json_case[i])
                    Testcase = []
                    for d in con_json_case[i].items():
                        Testcase.append(d[1])
                    mavdb.get_cursor(self)
                    Testcase = [str(var) for var in Testcase]
                    sql = '''
                        insert into faultcase
                        (CaseID,Subsystem,FaultType,ControlSequence,InterestedLog,TestStatus,UniqueIndex)
                        values(?,?,?,?,?,?,?)
                    '''
                    values = Testcase
                    # 如果用例库中存在CaseID主键冲突,则删除用例库中的，添加配置文件的
                    mavdb.is_TheSame_ID_pro(self,values[0])
                    self.cursor.execute(sql,values)
                    self.mydb.commit()
                
            # 6、将json_case中的剩余的case补进用例库
            #1)取出剩余的jsonID,求json_ID与con_ID的差集即可
            discon_jsonID = list(set(self.json_ID) - set(self.con_ID))
            #2)取出对应ID的case
            discon_json_case = []
            for i in range(len(json_case)):
                if json_case[i].get('CaseID') in discon_jsonID:
                    discon_json_case.append(json_case[i])
            #3）补到用例库
            for i in range(len(discon_json_case)):
                Testcase = []
                for d in  discon_json_case[i].items():
                    Testcase.append(d[1])
                mavdb.get_cursor(self)
                Testcase = [str(var) for var in Testcase]
                sql = '''
                    insert into faultcase
                    (CaseID,Subsystem,FaultType,ControlSequence,InterestedLog,TestStatus,UniqueIndex)
                    values(?,?,?,?,?,?,?)
                '''
                values = Testcase
                # 如果用例库中存在CaseID主键冲突,则删除用例库中的，添加配置文件的
                mavdb.is_TheSame_ID_pro(self,values[0])
                self.cursor.execute(sql,values)
                self.mydb.commit()

    # 更改json文件测试状态信息
    def json_pro(self,case_id):
        path = self.jsonpath + '/db.json'
        with open(path, "r",encoding='utf-8') as f:
            db_data = json.load(f)
        json_case = db_data.get('faultcase')

        if len(json_case) >= 1:
            # 如果CaseID存在于json文件中，则更改
            if case_id in self.json_ID:
                path = self.jsonpath + '/db.json'

                # 打开原json文件，读取内容并做修改
                with open(path, "r", encoding='utf-8') as f:
                    db_data = json.load(f)
                    # 更改测试状态
                    # db_data['faultcase'][case_id-1]['TestStatus'] = 'Finished'
                    data = db_data
                f.close()

                # 将更改后的内容写回json文件
                with open(path, "w",encoding='utf-8') as w:
                    json.dump(data,w,indent=4) # indent=4会自动换行，否则会变成1行
                w.close()


class DataAPI:
    def __init__(self, CaseID, UniqueIndex):
        self.jsonpath = sys.path[0] + '/db.json'

        # 获取平台log路径
        with open(self.jsonpath, "r", encoding='utf-8') as f:
            db_data = json.load(f)
        PlatFormpath = db_data.get('PlatFormPath')  # 为字典，需将字典的键提取出来形成字符串列表
        log_path = PlatFormpath + '/Firmware/build/px4_sitl_default/instance_1/log'
        PlatForm_log_dirs = os.listdir(log_path)  # 列出文件下的目录
        log_data = PlatForm_log_dirs[len(PlatForm_log_dirs)-1]
        self.path = os.path.join(log_path, log_data)  # ulg文件绝对路径

        self.ProjectPath = os.path.dirname(os.path.dirname(__file__))
        self.DataPath = self.ProjectPath + '/Data'
        self.CaseID = CaseID
        self.UniqueIndex = UniqueIndex
        self.TargetFilder = self.DataPath + '/TestCase_{}_{}'.format(self.CaseID, self.UniqueIndex)  # 获取当前测试用例的ID和目标文件夹
        self.TargetFilder_log = self.TargetFilder + '/Log'  # log文件夹路径
        self.TargetFilder_Tlog = self.TargetFilder + '/TLog'  # Tlog文件夹路径
        self.TargetFilder_truedata = self.TargetFilder + '/TrueData'  # 用户选择的文件夹路径
        self.TargetFilder_RateModelData = self.TargetFilder + '/RataModelData'  # 实时率模安全得分数据文件夹
        DataAPI.FolderCreat(self)

    def FolderCreat(self):
        # 生成指定文件夹（一个是log文件夹，一个是truedata文件夹，一个是RateModelData文件夹。）
        if os.path.exists(self.TargetFilder):  # 如果目标文件夹存在，则删除重建
            shutil.rmtree(self.TargetFilder)

        os.makedirs(self.TargetFilder)  # 创建TargetFilder目标文件夹---Testcase*文件夹
        os.makedirs(self.TargetFilder_log)  # 创建TargetFilder_log目标文件夹---log文件夹
        os.makedirs(self.TargetFilder_Tlog)  # 创建TargetFilder_Tlog目标文件夹---Tlog文件夹
        os.makedirs(self.TargetFilder_truedata)  # 创建TargetFilder_truedata目标文件夹---truedata文件夹


    def TLogDataExtra(self):
        # 1、列出文件下的目录
        Tlog_path = 'C:/Users/DELL/Documents/QGroundControl/Telemetry'
        listOfTlogFiles = [f for f in os.listdir(Tlog_path) if f[-5:] == ".tlog"]	# get list of only tlog files in current dir.
        listOfCSVFiles = [f for f in os.listdir(Tlog_path) if f[-4:] == ".csv"]	 # get list of only csv files in current dir.

        # 2、获取最新的tlg文件  即目录最下面那个tlg
        tlg = listOfTlogFiles[len(listOfTlogFiles)-1]
        tlgPath = os.path.join(Tlog_path, tlg)  # tlg文件绝对路径

        f_csv = listOfCSVFiles[len(listOfCSVFiles)-1]
        csvPath = os.path.join(Tlog_path, f_csv)  # csv文件绝对路径

        # 3、移动ulg文件到log文件夹
        TargetPath_Tlog = self.TargetFilder_Tlog + '/{}'.format(tlg)
        shutil.move(tlgPath, TargetPath_Tlog)  # 移动文件

        TargetPath_CSV = self.TargetFilder_Tlog + '/{}'.format(f_csv)
        shutil.move(csvPath, TargetPath_CSV)  # 移动文件



    def LogDataExtra(self):  # 提取log文件，并解析出感兴趣的csv文件
        # 1、列出文件下的目录
        dirs = os.listdir(self.path)  # 列出文件下的目录

        # 2、获取最新的ulg文件  即目录最下面那个ulg
        ulg = dirs[len(dirs)-1]
        ulgPath = os.path.join(self.path, ulg)  # ulg文件绝对路径

        # 3、复制ulg文件到log文件夹
        TargetPath_log = self.TargetFilder_log + '/{}'.format(ulg)
        shutil.copyfile(ulgPath, TargetPath_log)  # 复制文件

        # 4、将ulg转换成csv
        os.chdir(self.TargetFilder_log)
        os.system("for %i in (*); do ulog2csv %i")

        # 5、保留ulg文件,删除不要的csv文件
        # 1)加载json，获取感兴趣的csv文件
        with open(self.jsonpath, "r", encoding='utf-8') as f:
            db_data = json.load(f)

        csvfiles = db_data.get('testdata')  # 为字典，需将字典的键提取出来形成字符串列表
        csvfiles_keys = []
        for key in csvfiles.keys():
            csvfiles_keys.append(key)

        # 2)文件处理
        csvdirs = os.listdir(self.TargetFilder_log)
        csvdirs.remove(csvdirs[0])  # 保留ulg文件
        for file in csvdirs:
            flag = 0
            for i in range(len(csvfiles_keys)):
                if(fnmatch(file,csvfiles_keys[i])):
                    flag = 1
            if flag == 0: # 不感兴趣的csv文件
                fpath = os.path.join(self.TargetFilder_log,file) # 路径组合（获取csv的绝对路径）
                os.remove(fpath) # 删除文件

        # 3)删除多余解析的文件yaw_estimator_status_0
        remaindirs = os.listdir(self.TargetFilder_log)
        for file in remaindirs:
            if(fnmatch(file,'*yaw_estimator_status_0.csv')):
                os.remove(file) # 删除文件

    def TruedataRecord(self,mean_arr_v,mean_arr_ang,mean_arr_pos):
        # 创建表格
        wb = Workbook()
        # 创建表单
        sheet1 = wb.create_sheet('v')
        sheet2 = wb.create_sheet('ang')
        sheet3 = wb.create_sheet('pos')
        # 操作表单
        for i in range(len(mean_arr_v)):
            # 单个数直接用[]转化成列表,多个数用list强制转化为list类型
            v = list(mean_arr_v[i])
            sheet1.append(v)
            ang = list(mean_arr_ang[i])
            sheet2.append(ang)
            pos = list(mean_arr_pos[i])
            sheet3.append(pos)

        truedata_xlsxpath = self.TargetFilder_truedata + '//truedata.xlsx'
        wb.save(truedata_xlsxpath)
        # xlsx_to_csv
        data_xls_v = pd.read_excel(truedata_xlsxpath,sheet_name='v')
        truedata_csvpath_v = self.TargetFilder_truedata + '//truedata_v.csv' 
        data_xls_v.to_csv(truedata_csvpath_v, encoding='utf-8',index=False,header=None) # 转化为csv,index=False即为写入csv是不添加行索引，header=None为不添加每列的标题

        data_xls_ang = pd.read_excel(truedata_xlsxpath,sheet_name='ang')
        truedata_csvpath_ang = self.TargetFilder_truedata + '//truedata_ang.csv' 
        data_xls_ang.to_csv(truedata_csvpath_ang, encoding='utf-8',index=False,header=None) # 转化为csv

        data_xls_pos = pd.read_excel(truedata_xlsxpath,sheet_name='pos')
        truedata_csvpath_pos = self.TargetFilder_truedata + '//truedata_pos.csv' 
        data_xls_pos.to_csv(truedata_csvpath_pos, encoding='utf-8',index=False,header=None) # 转化为csv
        os.remove(truedata_xlsxpath) # 删除原来的xlsx文件

    def TestInformationRecord(self, TestInfo):
        # 创建表格
        wb = Workbook()
        # 创建表单
        sheet1 = wb.active
        sheet1.title = "TestInformation"
        # sheet1 = wb.create_sheet('TestInformation')
        # 操作表单
        for i in range(len(TestInfo)):
            # 单个数直接用[]转化成列表,多个数用list强制转化为list类型
            Testinfo = list(TestInfo[i])
            sheet1.append(Testinfo)
        
        TestInfo_xlsx_path = self.TargetFilder + '//TestInfo.xlsx'
        wb.save(TestInfo_xlsx_path)
        # xlsx_to_csv
        TestInfoData = pd.read_excel(TestInfo_xlsx_path, sheet_name='TestInformation')
        TestInfoData_csv_path = self.TargetFilder + '//TestInfo.csv'
        TestInfoData.to_csv(TestInfoData_csv_path, encoding='utf-8', index=False, index_label=False) # 转化为csv
        os.remove(TestInfo_xlsx_path) # 删除原来的xlsx文件
    

        



    