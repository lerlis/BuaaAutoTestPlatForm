# Copyright (C) RflySim from Central South University and
# rflylab from School of Automation Science and Electrical Engineering, Beihang University.
# All Rights Reserved.
from ast import While
from random import randint, random
from tokenize import group
from chardet import detect_all
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import PX4MavCtrlV4 as PX4MavCtrl
import matplotlib.pyplot as plt
from matplotlib.pyplot import MultipleLocator
import math
import re

# 记录差值序列数据
def fly_log_record_allan(Fault_Allan_deta_acc_group,Fault_Allan_deta_ang_group,Fault_Allan_deta_v_group,Fault_Allan_deta_pos_group):
    # 创建表格
    wb = Workbook()
    # 创建表单
    sheet1 = wb.create_sheet('deta_acc')
    sheet2 = wb.create_sheet('deta_ang')
    sheet3 = wb.create_sheet('deta_v')
    sheet4 = wb.create_sheet('deta_pos')
    # 操作表单
    for i in range(len(Fault_Allan_deta_acc_group)):
        # 单个数直接用[]转化成列表
        a = [Fault_Allan_deta_acc_group[i]]
        sheet1.append(a)
        b = [Fault_Allan_deta_ang_group[i]]
        sheet2.append(b)
        c = [Fault_Allan_deta_v_group[i]]
        sheet3.append(c)
        d = [Fault_Allan_deta_pos_group[i]]
        sheet4.append(d)

    wb.save('C://Users//dzl20201126//Desktop//差值序列.xlsx')

# 记录率模得分数据
def Rate_score_record(ang_score_group_Rate_Model_group,v_score_group_Rate_Model_group,pos_score_group_Rate_Model_group,Sys_Rate_Model_Helth_group):
    # 创建表格
    wb = Workbook()
    # 创建表单
    sheet5 = wb.create_sheet('Rate_ang')
    sheet6 = wb.create_sheet('Rate_v')
    sheet7 = wb.create_sheet('Rate_pos')
    sheet8 = wb.create_sheet('Rate_sys')
    # 操作表单
    for i in range(len(ang_score_group_Rate_Model_group)):
        # 单个数直接用[]转化成列表
        a = [ang_score_group_Rate_Model_group[i]]
        sheet5.append(a)
        b = [v_score_group_Rate_Model_group[i]]
        sheet6.append(b)
        c = [pos_score_group_Rate_Model_group[i]]
        sheet7.append(c)
        d = [Sys_Rate_Model_Helth_group[i]]
        sheet8.append(d)

    wb.save('C://Users//dzl20201126//Desktop//率模得分数据.xlsx')

# 整段飞行Allan方差分析
def Allan_analysis(mean_arr_ang,mean_arr_acc,mean_arr_v,mean_arr_pos):
    # 数据预处理：将三维变成一维(v和pos不用变一维)
    mean_arr_ang_sum = np.array([])
    mean_arr_acc_sum = np.array([])
    for i in range(len(mean_arr_ang)):
        mean_arr_ang_sum = np.append(mean_arr_ang_sum,np.sum(mean_arr_ang[i]))
        mean_arr_acc_sum = np.append(mean_arr_acc_sum,np.sum(mean_arr_acc[i]))
    # 方差分析
    # 1、分组
    Allan_group_num = 2 # 每组两个数据
    # Allan_group_num = one_num # 每组1s个数据
    Allan_mean_acc_group = np.array([])
    Allan_mean_ang_group = np.array([])
    Allan_mean_v_group = np.array([])
    Allan_mean_pos_group = np.array([])
    # 2、求均值序列
    i = 0 
    while True:
        if i >= len(mean_arr_acc_sum):
            break

        mean_acc = np.mean(mean_arr_acc_sum[i:i+Allan_group_num])
        mean_ang = np.mean(mean_arr_ang_sum[i:i+Allan_group_num])
        mean_v = sum(mean_arr_v[i:i+Allan_group_num])/Allan_group_num
        mean_pos = sum(mean_arr_pos[i:i+Allan_group_num])/Allan_group_num
        Allan_mean_acc_group = np.append(Allan_mean_acc_group,mean_acc)
        Allan_mean_ang_group = np.append(Allan_mean_ang_group,mean_ang)
        Allan_mean_v_group = np.append(Allan_mean_v_group,mean_v)
        Allan_mean_pos_group = np.append(Allan_mean_pos_group,mean_pos)
        i = i + Allan_group_num
    # 将vel、pos转化为三列
    Allan_mean_v_group = np.reshape(Allan_mean_v_group,(-1,3))
    Allan_mean_pos_group = np.reshape(Allan_mean_pos_group,(-1,3))


    exp_v = [0,0,0]
    exp_pos = [0,0,0]
    # 3、求差值序列
    j = 1
    Allan_deta_acc_group = np.array([])
    Allan_deta_ang_group = np.array([])
    Allan_deta_v_group = np.array([])
    Allan_deta_pos_group = np.array([])
    while True:
        if j >= len(Allan_mean_acc_group):
            break

        deta_acc = (Allan_mean_acc_group[j] - Allan_mean_acc_group[j-1])**2
        deta_ang = (Allan_mean_ang_group[j] - Allan_mean_ang_group[j-1])**2
        deta_v = (Allan_mean_v_group[j] - exp_v)
        deta_pos = (Allan_mean_pos_group[j] - exp_pos)
        Allan_deta_acc_group = np.append(Allan_deta_acc_group,deta_acc)
        Allan_deta_ang_group = np.append(Allan_deta_ang_group,deta_ang)
        Allan_deta_v_group = np.append(Allan_deta_v_group,deta_v)
        Allan_deta_pos_group = np.append(Allan_deta_pos_group,deta_pos)
        j = j + 1
    # 将vel、pos转化为三列
    Allan_deta_v_group = np.reshape(Allan_deta_v_group,(-1,3))
    Allan_deta_pos_group = np.reshape(Allan_deta_pos_group,(-1,3))
    Allan_deta_pos_group_sum = np.array([])
    Allan_deta_v_group_sum = np.array([])
    for i in range(len(Allan_deta_v_group)):
        Allan_deta_v_group_sum = np.append(Allan_deta_v_group_sum,np.sum(Allan_deta_v_group[i]))
        Allan_deta_pos_group_sum = np.append(Allan_deta_pos_group_sum,np.sum(Allan_deta_pos_group[i]))

    # 4、求Allan方差
    Allan_var_acc = 0
    Allan_var_ang = 0
    sum_acc = 0
    sum_ang = 0
    for i in range(len(Allan_deta_acc_group)):
        sum_acc = sum_acc + Allan_deta_acc_group[i]
        sum_ang = sum_ang + Allan_deta_ang_group[i]
    Allan_var_acc = round((1/(2*len(Allan_deta_acc_group)))*sum_acc,2)
    Allan_var_ang = round((1/(2*len(Allan_deta_ang_group)))*sum_ang,2)

    # Rate_Model_rank(Allan_deta_acc_group,Allan_deta_ang_group,Allan_deta_v_group_sum,Allan_deta_pos_group_sum)

    # 画图分析其均值差
    x = np.linspace(1,len(Allan_deta_acc_group),len(Allan_deta_acc_group))
    plt.figure(1)
    plt.subplot(1,2,1) # 画子图,一行5列子图,画在第3个图上
    plt.xlabel("ang")
    plt.ylabel('deta')
    plt.plot(x,Allan_deta_ang_group)
    plt.subplot(1,2,2) # 画子图,一行5列子图,画在第4个图上
    plt.xlabel("acc")
    plt.ylabel('deta')
    plt.plot(x,Allan_deta_acc_group)
    plt.show()

# 故障注入Allan方差分析
def Fault_inject_Allan(start_index,end_index,fault_inject_index,land_index,mean_arr_ang,mean_arr_v,one_num,cmd):
    global Fault_Allan_deta_ang_group
    # 方差分析
    # 数据预处理
    Fault_mean_arr_ang = mean_arr_ang[start_index:end_index]
    Fault_mean_arr_v = mean_arr_v[start_index:end_index]
    # 三维变一维(vel、pos不用变一维)
    Fault_mean_arr_ang_sum = np.array([])
  
    for i in range(len(Fault_mean_arr_ang)):
        Fault_mean_arr_ang_sum = np.append(Fault_mean_arr_ang_sum,np.sum(Fault_mean_arr_ang[i]))
        
    # 1、分组
    Fault_Allan_group_num = 2 # 每组两个数据
    # Fault_Allan_group_num = one_num # 每组1s个数据
    Fault_Allan_mean_ang_group = np.array([])
    Fault_Allan_mean_v_group = np.array([])

    # 2、求均值序列(其中Fault_Allan_mean_pos_group需转化为三列)
    i = 0 
    while True:
        if i >= len(Fault_mean_arr_ang_sum):
            break

        mean_ang = np.mean(Fault_mean_arr_ang_sum[i:i+Fault_Allan_group_num])
        mean_v = sum(Fault_mean_arr_v[i:i+Fault_Allan_group_num])/Fault_Allan_group_num
        Fault_Allan_mean_ang_group = np.append(Fault_Allan_mean_ang_group,mean_ang)
        Fault_Allan_mean_v_group = np.append(Fault_Allan_mean_v_group,mean_v)
        i = i + Fault_Allan_group_num
    # 将vel、pos转化为三列
    Fault_Allan_mean_v_group = np.reshape(Fault_Allan_mean_v_group,(-1,3))

    # 判断是否有速度控制指令
    isVelCtrl = 0
    for i in range(len(cmd)):
        pro = re.findall(r'-?\d+\.?[0-9]*',cmd[i])
        if pro[1] == '4':
            isVelCtrl = 1

    # 不使用位置指标
    exp_v = np.zeros_like(Fault_Allan_mean_v_group)


    # 如果没有指定速度指令，则使用默认的速度（即故障注入时刻的速度）
    if isVelCtrl == 0:
        desired_vx = Fault_Allan_mean_v_group[round((fault_inject_index-start_index)/Fault_Allan_group_num)][0]
        desired_vy = Fault_Allan_mean_v_group[round((fault_inject_index-start_index)/Fault_Allan_group_num)][1]
        desired_vz = Fault_Allan_mean_v_group[round((fault_inject_index-start_index)/Fault_Allan_group_num)][2]
        exp_v[:,:]= [desired_vx,desired_vy,desired_vz]
    # 如果使用了速度指令，则使用指定的速度
    elif isVelCtrl == 1:
        # 提取预期速度
        for i in range(len(cmd)):
            pro = re.findall(r'-?\d+\.?[0-9]*',cmd[i])
            if pro[1] == '4': # 4为速度控制指令
                vel = [pro[2],pro[3],pro[4]]
        vel = [float(var) for var in vel]
        exp_v[:,0] = vel[0]
        exp_v[:,1] = vel[1]
        exp_v[:,2] = vel[2]

    # 判断是否有降落控制指令
    isLandCtrl = 0
    for i in range(len(cmd)):
        pro = re.findall(r'-?\d+\.?[0-9]*',cmd[i])
        if pro[1] == '5': # 5为降落指令
            isLandCtrl = 1

    # 如果有降落标志，则期望速度应为降落时的速度
    if isLandCtrl == 1:
        # 降落速度默认为[0,0,-2]
        # sendvel的z轴上的不准确？---------------------------------------------------------------------------------
        exp_v[round((land_index-start_index)/Fault_Allan_group_num):,0] = 0
        exp_v[round((land_index-start_index)/Fault_Allan_group_num):,1] = 0
        exp_v[round((land_index-start_index)/Fault_Allan_group_num):,2] = -1
        

    global isFaultInject
    # 如果没有故障注入，则只使用ang一个指标
    isFaultInject = 0
    for i in range(len(cmd)):
        pro = re.findall(r'-?\d+\.?[0-9]*',cmd[i])
        if pro[1] == '6': # 6为故障注入指令
            isFaultInject = 1

    
    if isFaultInject == 0:
        # 不使用位置和速度指标
        exp_v = Fault_Allan_mean_v_group

    '''
    # 记录期望的位置和速度数据
    # 创建表格
    wb = Workbook()
    # 创建表单
    sheet9 = wb.create_sheet('exp_v')
    sheet10 = wb.create_sheet('exp_pos')
    # 操作表单
    for i in range(len(exp_v)):
        # 数组用转化成列表
        a = list(exp_v[i])
        sheet9.append(a)
        b = list(exp_pos[i])
        sheet10.append(b)

    wb.save('C://Users//dzl20201126//Desktop//期望位置速度.xlsx')
    '''
        
    # 3、求差值序列
    j = 1
    Fault_Allan_deta_ang_group = np.array([])
    Fault_Allan_deta_v_group = np.array([])
    while True:
        if j >= len(Fault_Allan_mean_ang_group):
            break

        deta_ang = (Fault_Allan_mean_ang_group[j] - Fault_Allan_mean_ang_group[j-1])**2
        deta_v = abs((Fault_Allan_mean_v_group[j] - exp_v[j]))
        Fault_Allan_deta_ang_group = np.append(Fault_Allan_deta_ang_group,deta_ang)
        Fault_Allan_deta_v_group = np.append(Fault_Allan_deta_v_group,deta_v)
        j = j + 1
    # 将vel、pos转化为三列,并每列求和求出每组之间的和
    Fault_Allan_deta_v_group = np.reshape(Fault_Allan_deta_v_group,(-1,3))
    Fault_Allan_deta_v_group_sum = np.array([])
    for i in range(len(Fault_Allan_deta_v_group)):
        Fault_Allan_deta_v_group_sum = np.append(Fault_Allan_deta_v_group_sum,np.sum(Fault_Allan_deta_v_group[i]))

    # print('差值序列长度:',len(Fault_Allan_deta_acc_group))

    # 4、求Allan方差
    Fault_Allan_var_ang = 0
    Fault_sum_ang = 0
    for i in range(len(Fault_Allan_deta_ang_group)):
        Fault_sum_ang = Fault_sum_ang + Fault_Allan_deta_ang_group[i]
    Fault_Allan_var_ang = round((1/(2*len(Fault_Allan_deta_ang_group)))*Fault_sum_ang,2)

    # print('acc波动均值',np.mean(Fault_Allan_deta_acc_group))
    # print('ang波动均值',np.mean(Fault_Allan_deta_ang_group))
    # print('v波动均值',np.mean(Fault_Allan_deta_v_group_sum))
    # print('pos波动均值',np.mean(Fault_Allan_deta_pos_group_sum))

    # 记录差值序列数据
    # fly_log_record_allan(Fault_Allan_deta_acc_group,Fault_Allan_deta_ang_group,Fault_Allan_deta_v_group_sum,Fault_Allan_deta_pos_group_sum)
    # 记录率模得分数据
    Rate_Model_rank(Fault_Allan_deta_ang_group,Fault_Allan_deta_v_group_sum,one_num)
    '''
    # 画图分析其均值差
    x = np.linspace(1,len(Fault_Allan_deta_acc_group),len(Fault_Allan_deta_acc_group))
    plt.figure(1)
    plt.subplot(1,4,1) # 画子图,一行5列子图,画在第3个图上
    plt.xlabel("ang")
    plt.ylabel('deta')
    plt.plot(x,Fault_Allan_deta_ang_group)
    plt.subplot(1,4,2) # 画子图,一行5列子图,画在第4个图上
    plt.xlabel("acc")
    plt.ylabel('deta')
    plt.plot(x,Fault_Allan_deta_acc_group)
    plt.subplot(1,4,3) # 画子图,一行5列子图,画在第4个图上
    plt.xlabel("v")
    plt.ylabel('deta')
    plt.plot(x,Fault_Allan_deta_v_group_sum)
    plt.subplot(1,4,4) # 画子图,一行5列子图,画在第4个图上
    plt.xlabel("pos")
    plt.ylabel('deta')
    plt.plot(x,Fault_Allan_deta_pos_group_sum)
    plt.show()
    '''
    

# 画每隔x秒的波动范围图
def range_bodong(Fault_Allan_deta_ang_group,Fault_Allan_deta_v_group_sum,Fault_Allan_deta_pos_group_sum,one_num):
    # 每0.2s收集一组数据
    num = round(one_num/5)  
    num_deta_ang_group = np.array([])
    num_deta_v_group = np.array([])
    num_deta_pos_group = np.array([])
    i = 0
    while True:
        # 收集最后一组数据
        if i >= len(Fault_Allan_deta_ang_group):
            i = len(Fault_Allan_deta_ang_group)-1
            num_deta_ang_group = np.append(num_deta_ang_group,Fault_Allan_deta_ang_group[i])
            num_deta_v_group = np.append(num_deta_v_group,Fault_Allan_deta_v_group_sum[i])
            num_deta_pos_group = np.append(num_deta_pos_group,Fault_Allan_deta_pos_group_sum[i])
            break

        num_deta_ang_group = np.append(num_deta_ang_group,Fault_Allan_deta_ang_group[i])
        num_deta_v_group = np.append(num_deta_v_group,Fault_Allan_deta_v_group_sum[i])
        num_deta_pos_group = np.append(num_deta_pos_group,Fault_Allan_deta_pos_group_sum[i])
        i = i + num
    

    # print('0.2s的ang波动均值',np.mean(num_deta_ang_group))
    # print('0.2s的v波动均值',np.mean(num_deta_v_group))
    # print('0.2s的pos波动均值',np.mean(num_deta_pos_group))

    x = np.linspace(1,len(num_deta_ang_group),len(num_deta_ang_group))
    plt.subplot(131)
    plt.plot(x,num_deta_ang_group)
    plt.xlabel('ang')
    plt.subplot(132)
    plt.plot(x,num_deta_v_group)
    plt.xlabel('v')
    plt.subplot(133)
    plt.plot(x,num_deta_pos_group)
    plt.xlabel('pos')
    plt.show()

# 未坠机评估：飞行健康评估:率模健康度
def Rate_Model_rank(Fault_Allan_deta_ang_group,Fault_Allan_deta_v_group_sum,one_num):
    global sys_weight_score,Rate_rank
    # 0.2s的常值波动范围
    Allan_ang_range_Rate_Model = {
        '正常':[0,0.0005],     # (1,0.9)
        '轻微':[0.0005,0.005], # (0.9,0.7)
        '严重':[0.005,0.025],  # (0.7,0.4)
        '危险':[0.025,0.25],    # (0.4,0.1)
    }
    Allan_v_range_Rate_Model = {
        '正常':[0,0.1],     # (1,0.9)
        '轻微':[0.1,0.35], # (0.9,0.7)
        '严重':[0.35,1.5],  # (0.7,0.4)
        '危险':[1.5,2],    # (0.4,0.1)
    }
    rank_range = {
        '正常':[0.9,0.999],
        '轻微':[0.7,0.9],       
        '严重':[0.4,0.7],         
        '危险':[0.1,0.4],    
    }
    '''
    # 原始的波动范围，非均值波动
    Allan_ang_range_Rate_Model = {
        '正常':[0,0.002],     # (1,0.9)
        '轻微':[0.002,0.025], # (0.9,0.7)
        '严重':[0.025,0.45],  # (0.7,0.4)
        '危险':[0.45,2],    # (0.4,0.1)
    }
    Allan_v_range_Rate_Model = {
        '正常':[0,0.1],     # (1,0.9)
        '轻微':[0.1,0.35], # (0.9,0.7)
        '严重':[0.35,1.5],  # (0.7,0.4)
        '危险':[1.5,4],    # (0.4,0.1)
    }
    Allan_pos_range_Rate_Model = {
        '正常':[0,0.2],     # (1,0.9)
        '轻微':[0.2,0.5], # (0.9,0.7)
        '严重':[0.5,2],  # (0.7,0.4)
        '危险':[2,3],    # (0.4,0.1)
    }
    rank_range = {
        '正常':[0.9,0.999],
        '轻微':[0.7,0.9],       
        '严重':[0.4,0.7],         
        '危险':[0.1,0.4],    
    }
    '''
    # 每0.2s评价一次风险
    one_num_Allan = one_num/2
    num = round(one_num_Allan/5)
    # 权重参数
    weight_nomal = 0.1
    weight_slight = 0.3
    weight_serious = 0.9
    weight_hazard = 1.5
    weight_disaster = 3

    # 模糊综合评价参数
    Tao_matrix = np.zeros([2,5])
    
    global ang_weight_score,v_weight_score,sys_weight_score,Rate_rank
    # 求不同健康状态的衰减模型参数k和a
    # 正常
    k_nomal_ang = (Allan_ang_range_Rate_Model.get('正常')[0]*math.log(rank_range.get('正常')[0]) - Allan_ang_range_Rate_Model.get('正常')[1]*math.log(rank_range.get('正常')[1]))/(math.log(rank_range.get('正常')[1])-math.log(rank_range.get('正常')[0]))
    a_nomal_ang = (-math.log(rank_range.get('正常')[1]))/(Allan_ang_range_Rate_Model.get('正常')[0]+k_nomal_ang)
    # 轻微
    k_slight_ang = (Allan_ang_range_Rate_Model.get('轻微')[0]*math.log(rank_range.get('轻微')[0]) - Allan_ang_range_Rate_Model.get('轻微')[1]*math.log(rank_range.get('轻微')[1]))/(math.log(rank_range.get('轻微')[1])-math.log(rank_range.get('轻微')[0]))
    a_slight_ang = (-math.log(rank_range.get('轻微')[1]))/(Allan_ang_range_Rate_Model.get('轻微')[0]+k_slight_ang)
    # 严重
    k_serious_ang = (Allan_ang_range_Rate_Model.get('严重')[0]*math.log(rank_range.get('严重')[0]) - Allan_ang_range_Rate_Model.get('严重')[1]*math.log(rank_range.get('严重')[1]))/(math.log(rank_range.get('严重')[1])-math.log(rank_range.get('严重')[0]))
    a_serious_ang = (-math.log(rank_range.get('严重')[1]))/(Allan_ang_range_Rate_Model.get('严重')[0]+k_serious_ang)
    # 危险
    k_hazard_ang = (Allan_ang_range_Rate_Model.get('危险')[0]*math.log(rank_range.get('危险')[0]) - Allan_ang_range_Rate_Model.get('危险')[1]*math.log(rank_range.get('危险')[1]))/(math.log(rank_range.get('危险')[1])-math.log(rank_range.get('危险')[0]))
    a_hazard_ang = (-math.log(rank_range.get('危险')[1]))/(Allan_ang_range_Rate_Model.get('危险')[0]+k_hazard_ang)

    i_ang = 0
    ang_score_group_Rate_Model = np.array([])
    while True:
        # 收集最后一组数据
        if i_ang + num >= len(Fault_Allan_deta_ang_group):
            i_ang = len(Fault_Allan_deta_ang_group) - 1
            if Allan_ang_range_Rate_Model.get('正常')[0] <= Fault_Allan_deta_ang_group[i_ang] < Allan_ang_range_Rate_Model.get('正常')[1]:
                ang_score = math.exp((-a_nomal_ang)*(Fault_Allan_deta_ang_group[i_ang]+k_nomal_ang))
                ang_score_group_Rate_Model = np.append(ang_score_group_Rate_Model,ang_score)
            elif Allan_ang_range_Rate_Model.get('轻微')[0] <= Fault_Allan_deta_ang_group[i_ang] < Allan_ang_range_Rate_Model.get('轻微')[1]:
                ang_score = math.exp((-a_slight_ang)*(Fault_Allan_deta_ang_group[i_ang]+k_slight_ang))
                ang_score_group_Rate_Model = np.append(ang_score_group_Rate_Model,ang_score)
            elif Allan_ang_range_Rate_Model.get('严重')[0] <= Fault_Allan_deta_ang_group[i_ang] < Allan_ang_range_Rate_Model.get('严重')[1]:
                ang_score = math.exp((-a_serious_ang)*(Fault_Allan_deta_ang_group[i_ang]+k_serious_ang))
                ang_score_group_Rate_Model = np.append(ang_score_group_Rate_Model,ang_score)
            elif Allan_ang_range_Rate_Model.get('危险')[0] <= Fault_Allan_deta_ang_group[i_ang] < Allan_ang_range_Rate_Model.get('危险')[1]:
                ang_score = math.exp((-a_hazard_ang)*(Fault_Allan_deta_ang_group[i_ang]+k_hazard_ang))
                ang_score_group_Rate_Model = np.append(ang_score_group_Rate_Model,ang_score)
            else:
                ang_score_group_Rate_Model = np.append(ang_score_group_Rate_Model,0.1)
            break

        if Allan_ang_range_Rate_Model.get('正常')[0] <= Fault_Allan_deta_ang_group[i_ang] < Allan_ang_range_Rate_Model.get('正常')[1]:
            ang_score = math.exp((-a_nomal_ang)*(Fault_Allan_deta_ang_group[i_ang]+k_nomal_ang))
            ang_score_group_Rate_Model = np.append(ang_score_group_Rate_Model,ang_score)
        elif Allan_ang_range_Rate_Model.get('轻微')[0] <= Fault_Allan_deta_ang_group[i_ang] < Allan_ang_range_Rate_Model.get('轻微')[1]:
            ang_score = math.exp((-a_slight_ang)*(Fault_Allan_deta_ang_group[i_ang]+k_slight_ang))
            ang_score_group_Rate_Model = np.append(ang_score_group_Rate_Model,ang_score)
        elif Allan_ang_range_Rate_Model.get('严重')[0] <= Fault_Allan_deta_ang_group[i_ang] < Allan_ang_range_Rate_Model.get('严重')[1]:
            ang_score = math.exp((-a_serious_ang)*(Fault_Allan_deta_ang_group[i_ang]+k_serious_ang))
            ang_score_group_Rate_Model = np.append(ang_score_group_Rate_Model,ang_score)
        elif Allan_ang_range_Rate_Model.get('危险')[0] <= Fault_Allan_deta_ang_group[i_ang] < Allan_ang_range_Rate_Model.get('危险')[1]:
            ang_score = math.exp((-a_hazard_ang)*(Fault_Allan_deta_ang_group[i_ang]+k_hazard_ang))
            ang_score_group_Rate_Model = np.append(ang_score_group_Rate_Model,ang_score)
        else:
            ang_score_group_Rate_Model = np.append(ang_score_group_Rate_Model,0.1)
        
        i_ang = i_ang + num

    # ang隶属度
    u_s_ang = np.array([])
    for i in range(len(ang_score_group_Rate_Model)):
        if ang_score_group_Rate_Model[i]>0.9:
            u = 1
        else:
            u = ang_score_group_Rate_Model[i]
        u_s_ang = np.append(u_s_ang,u)
    ang_score_group_Rate_Model_group = np.array([])
    ang_score_group_Rate_Model_group = np.append(ang_score_group_Rate_Model_group,u_s_ang[0]) # 把第一个加进去
    # ang隶属度转移概率
    for i in range(len(u_s_ang)-1):
        ang_score_begin = u_s_ang[i]
        ang_score_end = u_s_ang[i+1]
        if ang_score_end <= ang_score_begin: # 当后一个健康度小于等于前一个时，应当使用状态转移
            u_Tsf = ang_score_begin - ang_score_end
            ang_score = ang_score_begin*(1-u_Tsf)
        else: # 当后一个健康度大于前一个时，应当保持上一时刻的计算值，以减小波动
            ang_score = u_s_ang[i]
        ang_score_group_Rate_Model_group = np.append(ang_score_group_Rate_Model_group,ang_score)
    # 求ang_score_group_Rate_Model_group的加权率模得分
    # 记录weight_ang数组
    weight_ang = np.array([])
    for i in range(len(ang_score_group_Rate_Model_group)):
        if rank_range.get('正常')[0] < ang_score_group_Rate_Model_group[i] <= 1:
            weight_ang = np.append(weight_ang,weight_nomal)
            Tao_matrix[0][0] += 1 # 正常个数加一
        elif rank_range.get('轻微')[0] < ang_score_group_Rate_Model_group[i] <= rank_range.get('轻微')[1]:
            weight_ang = np.append(weight_ang,weight_slight)
            Tao_matrix[0][1] += 1 # 轻微个数加一
        elif rank_range.get('严重')[0] < ang_score_group_Rate_Model_group[i] <= rank_range.get('严重')[1]:
            weight_ang = np.append(weight_ang,weight_serious)
            Tao_matrix[0][2] += 1 # 严重个数加一
        elif rank_range.get('危险')[0] < ang_score_group_Rate_Model_group[i] <= rank_range.get('危险')[1]:
            weight_ang = np.append(weight_ang,weight_hazard)
            Tao_matrix[0][3] += 1 # 危险个数加一
        else:
            weight_ang = np.append(weight_ang,weight_disaster)
            Tao_matrix[0][4] += 1 # 灾难个数加一
    weight_ang = weight_ang/sum(weight_ang)
    # 求加权后的ang隶属度结果
    ang_weight_score = 0
    for i in range(len(ang_score_group_Rate_Model_group)):
        ang_weight_score = ang_weight_score + weight_ang[i]*ang_score_group_Rate_Model_group[i]
    # print('ang率模得分:',round(ang_weight_score,3))

    # v健康风险状态划分（求不同健康状态的衰减模型参数k和a)
    # 正常
    k_nomal_v = (Allan_v_range_Rate_Model.get('正常')[0]*math.log(rank_range.get('正常')[0]) - Allan_v_range_Rate_Model.get('正常')[1]*math.log(rank_range.get('正常')[1]))/(math.log(rank_range.get('正常')[1])-math.log(rank_range.get('正常')[0]))
    a_nomal_v = (-math.log(rank_range.get('正常')[1]))/(Allan_v_range_Rate_Model.get('正常')[0]+k_nomal_v)
    # 轻微
    k_slight_v = (Allan_v_range_Rate_Model.get('轻微')[0]*math.log(rank_range.get('轻微')[0]) - Allan_v_range_Rate_Model.get('轻微')[1]*math.log(rank_range.get('轻微')[1]))/(math.log(rank_range.get('轻微')[1])-math.log(rank_range.get('轻微')[0]))
    a_slight_v = (-math.log(rank_range.get('轻微')[1]))/(Allan_v_range_Rate_Model.get('轻微')[0]+k_slight_v)
    # 严重
    k_serious_v = (Allan_v_range_Rate_Model.get('严重')[0]*math.log(rank_range.get('严重')[0]) - Allan_v_range_Rate_Model.get('严重')[1]*math.log(rank_range.get('严重')[1]))/(math.log(rank_range.get('严重')[1])-math.log(rank_range.get('严重')[0]))
    a_serious_v = (-math.log(rank_range.get('严重')[1]))/(Allan_v_range_Rate_Model.get('严重')[0]+k_serious_v)
    # 危险
    k_hazard_v = (Allan_v_range_Rate_Model.get('危险')[0]*math.log(rank_range.get('危险')[0]) - Allan_v_range_Rate_Model.get('危险')[1]*math.log(rank_range.get('危险')[1]))/(math.log(rank_range.get('危险')[1])-math.log(rank_range.get('危险')[0]))
    a_hazard_v = (-math.log(rank_range.get('危险')[1]))/(Allan_v_range_Rate_Model.get('危险')[0]+k_hazard_v)

    i_v = 0
    v_score_group_Rate_Model = np.array([])
    while True:
        # 收集最后一组数据
        if i_v + num >= len(Fault_Allan_deta_v_group_sum):
            i_v = len(Fault_Allan_deta_v_group_sum) - 1
            if Allan_v_range_Rate_Model.get('正常')[0] <= Fault_Allan_deta_v_group_sum[i_v] < Allan_v_range_Rate_Model.get('正常')[1]:
                v_score = math.exp((-a_nomal_v)*(Fault_Allan_deta_v_group_sum[i_v]+k_nomal_v))
                v_score_group_Rate_Model = np.append(v_score_group_Rate_Model,v_score)
            elif Allan_v_range_Rate_Model.get('轻微')[0] <= Fault_Allan_deta_v_group_sum[i_v] < Allan_v_range_Rate_Model.get('轻微')[1]:
                v_score = math.exp((-a_slight_v)*(Fault_Allan_deta_v_group_sum[i_v]+k_slight_v))
                v_score_group_Rate_Model = np.append(v_score_group_Rate_Model,v_score)
            elif Allan_v_range_Rate_Model.get('严重')[0] <= Fault_Allan_deta_v_group_sum[i_v] < Allan_v_range_Rate_Model.get('严重')[1]:
                v_score = math.exp((-a_serious_v)*(Fault_Allan_deta_v_group_sum[i_v]+k_serious_v))
                v_score_group_Rate_Model = np.append(v_score_group_Rate_Model,v_score)
            elif Allan_v_range_Rate_Model.get('危险')[0] <= Fault_Allan_deta_v_group_sum[i_v] < Allan_v_range_Rate_Model.get('危险')[1]:
                v_score = math.exp((-a_hazard_v)*(Fault_Allan_deta_v_group_sum[i_v]+k_hazard_v))
                v_score_group_Rate_Model = np.append(v_score_group_Rate_Model,v_score)
            else:
                v_score_group_Rate_Model = np.append(v_score_group_Rate_Model,0.1)
            break

        if Allan_v_range_Rate_Model.get('正常')[0] <= Fault_Allan_deta_v_group_sum[i_v] < Allan_v_range_Rate_Model.get('正常')[1]:
            v_score = math.exp((-a_nomal_v)*(Fault_Allan_deta_v_group_sum[i_v]+k_nomal_v))
            v_score_group_Rate_Model = np.append(v_score_group_Rate_Model,v_score)
        elif Allan_v_range_Rate_Model.get('轻微')[0] <= Fault_Allan_deta_v_group_sum[i_v] < Allan_v_range_Rate_Model.get('轻微')[1]:
            v_score = math.exp((-a_slight_v)*(Fault_Allan_deta_v_group_sum[i_v]+k_slight_v))
            v_score_group_Rate_Model = np.append(v_score_group_Rate_Model,v_score)
        elif Allan_v_range_Rate_Model.get('严重')[0] <= Fault_Allan_deta_v_group_sum[i_v] < Allan_v_range_Rate_Model.get('严重')[1]:
            v_score = math.exp((-a_serious_v)*(Fault_Allan_deta_v_group_sum[i_v]+k_serious_v))
            v_score_group_Rate_Model = np.append(v_score_group_Rate_Model,v_score)
        elif Allan_v_range_Rate_Model.get('危险')[0] <= Fault_Allan_deta_v_group_sum[i_v] < Allan_v_range_Rate_Model.get('危险')[1]:
            v_score = math.exp((-a_hazard_v)*(Fault_Allan_deta_v_group_sum[i_v]+k_hazard_v))
            v_score_group_Rate_Model = np.append(v_score_group_Rate_Model,v_score)
        else:
            v_score_group_Rate_Model = np.append(v_score_group_Rate_Model,0.1)
        
        i_v = i_v + num

    # v隶属度
    u_s_v = np.array([])
    for i in range(len(v_score_group_Rate_Model)):
        if v_score_group_Rate_Model[i]>0.9:
            u = 1
        else:
            u = v_score_group_Rate_Model[i]
        u_s_v = np.append(u_s_v,u)
    v_score_group_Rate_Model_group = np.array([])
    v_score_group_Rate_Model_group = np.append(v_score_group_Rate_Model_group,u_s_v[0]) # 把第一个加进去
    # v隶属度转移概率
    for i in range(len(u_s_v)-1):
        v_score_begin = u_s_v[i]
        v_score_end = u_s_v[i+1]
        if v_score_end <= v_score_begin: # 当后一个健康度小于等于前一个时，应当使用状态转移
            u_Tsf = v_score_begin - v_score_end
            v_score = v_score_begin*(1-u_Tsf)
        else: # 当后一个健康度大于前一个时，应当保持上一时刻的计算值，以减小波动
            v_score = u_s_v[i]
        v_score_group_Rate_Model_group = np.append(v_score_group_Rate_Model_group,v_score)

    # 求v_score_group_Rate_Model_group的加权率模得分
    # 记录weight_v数组
    weight_v = np.array([])
    for i in range(len(v_score_group_Rate_Model_group)):
        if rank_range.get('正常')[0] < v_score_group_Rate_Model_group[i] <= 1:
            weight_v = np.append(weight_v,weight_nomal)
            Tao_matrix[1][0] += 1 # 正常个数加一
        elif rank_range.get('轻微')[0] < v_score_group_Rate_Model_group[i] <= rank_range.get('轻微')[1]:
            weight_v = np.append(weight_v,weight_slight)
            Tao_matrix[1][1] += 1 # 轻微个数加一
        elif rank_range.get('严重')[0] < v_score_group_Rate_Model_group[i] <= rank_range.get('严重')[1]:
            weight_v = np.append(weight_v,weight_serious)
            Tao_matrix[1][2] += 1 # 严重个数加一
        elif rank_range.get('危险')[0] < v_score_group_Rate_Model_group[i] <= rank_range.get('危险')[1]:
            weight_v = np.append(weight_v,weight_hazard)
            Tao_matrix[1][3] += 1 # 危险个数加一
        else:
            weight_v = np.append(weight_v,weight_disaster)
            Tao_matrix[1][4] += 1 # 灾难个数加一
    weight_v = weight_v/sum(weight_v)
    # 求加权后的v隶属度结果
    v_weight_score = 0
    for i in range(len(v_score_group_Rate_Model_group)):
        v_weight_score = v_weight_score + weight_v[i]*v_score_group_Rate_Model_group[i]
    # print('v率模得分:',round(v_weight_score,3))

    # 归一化健康评判矩阵
    Wr_sum = np.sum(Tao_matrix,axis=1)[0] # axis=1 表示按每一行求和
    Dv_sum = np.sum(Tao_matrix,axis=1)[1]
    for i in range(Tao_matrix.shape[1]):
        Tao_matrix[0][i] = Tao_matrix[0][i]/Wr_sum
        Tao_matrix[1][i] = Tao_matrix[1][i]/Dv_sum

    # print('Tao_matrix',Tao_matrix)

    # 权重矩阵
    k_ang = 0.4
    k_v = 0.6
    W = np.array([k_ang,k_v])

    # 输出评判矩阵
    Health_Matrix = np.dot(W,Tao_matrix) # 注意是点乘，即矩阵乘法
    # print('Health_Matrix',Health_Matrix)

    # 键值向量
    Health_Level = {
        'Unaffected':Health_Matrix[0],
        'Slight':Health_Matrix[1],
        'Serious':Health_Matrix[2],
        'Dangerous':Health_Matrix[3],
        'Disaster':Health_Matrix[4]
    }

    # 输出最大值对应的键
    Health_UAV = max(Health_Level,key=Health_Level.get)
    Rate_rank = Health_UAV
    print('Sys Health Level:',Health_UAV)

    # 输出最终得分(法二：用所有指标的加权结果乘对应的权重输出)
    Score_UAV = round(k_ang * ang_weight_score + k_v * v_weight_score,3)
    sys_weight_score = Score_UAV
    print('Sys rate modulus score:',Score_UAV)

    '''
    # 系统率模得分
    # 法一:综合所有指标再加权sys
    Sys_Rate_Model_Helth_group = np.array([])
    for i in range(len(ang_score_group_Rate_Model_group)):
        all_score = k_ang*ang_score_group_Rate_Model_group[i]+k_v*v_score_group_Rate_Model_group[i]
        Sys_Rate_Model_Helth_group = np.append(Sys_Rate_Model_Helth_group,all_score)

    # 求sys_score_group_Rate_Model_group的加权率模得分
    # 记录weight_sys数组
    weight_sys = np.array([])
    for i in range(len(Sys_Rate_Model_Helth_group)):
        if rank_range.get('正常')[0] < Sys_Rate_Model_Helth_group[i] <= 1:
            weight_sys = np.append(weight_sys,weight_nomal)
        elif rank_range.get('轻微')[0] < Sys_Rate_Model_Helth_group[i] <= rank_range.get('轻微')[1]:
            weight_sys = np.append(weight_sys,weight_slight)
        elif rank_range.get('严重')[0] < Sys_Rate_Model_Helth_group[i] <= rank_range.get('严重')[1]:
            weight_sys = np.append(weight_sys,weight_serious)
        elif rank_range.get('危险')[0] < Sys_Rate_Model_Helth_group[i] <= rank_range.get('危险')[1]:
            weight_sys = np.append(weight_sys,weight_hazard)
        else:
            weight_sys = np.append(weight_sys,weight_hazard)
    weight_sys = weight_sys/sum(weight_sys)
    # 求加权后的sys隶属度结果
    sys_weight_score = 0
    for i in range(len(Sys_Rate_Model_Helth_group)):
        sys_weight_score = sys_weight_score + weight_sys[i]*Sys_Rate_Model_Helth_group[i]
    print('sys率模得分:',round(sys_weight_score,3))   
    '''

# 坠机评估
def Fall_rank(fall_energy):
    '''
    坠机等级范围:
    可接受:[0.07,0.1)
    严重:[0.025,0.07)
    灾难:[0,0.025]
    '''
    global Sys_score,Sys_tag
    alpha = 10**6
    beta = 100
    E_imp = fall_energy
    Ps = 0.1 
    P = 1/(1+((alpha/beta)*(beta/E_imp)**(1/(4*Ps)))**0.5)

    R1 = 10**(-8)
    k = 5*10**(-6)
    R = (math.exp((P/R1)*k)-1)/(math.exp((P/R1)*k)+1)

    Sys_score = round(0.1-(R/10),3)
    Sys_tag = 0
    if 0 <= Sys_score < 0.025:
        Sys_tag = 'Crash: Disaster'
    elif 0.025 <= Sys_score < 0.07:
        Sys_tag = 'Crash: Severe'
    else:
        Sys_tag = 'Crash: Mild'

    # print('坠地严重程度:',P)
    # print('坠地指标风险:',R)
    print('Fall risk',Sys_score)
    print('Fall level:',Sys_tag)

def draw_pict(ang_score_group_Rate_Model_group,v_score_group_Rate_Model_group,Sys_Rate_Model_Helth_group,record_time,break_time):
    x1 = np.linspace(record_time-record_time,break_time-record_time,len(ang_score_group_Rate_Model_group))
    y_major_locator=MultipleLocator(0.5)#以每0.5显示
    ax=plt.gca()
    plt.figure(1)
    plt.plot(x1,ang_score_group_Rate_Model_group)
    plt.xlabel('Time(s)')
    plt.ylabel('W_r')
    ax.yaxis.set_major_locator(y_major_locator)
    plt.ylim(0,1.5)
    plt.figure(2)
    plt.plot(x1,v_score_group_Rate_Model_group)
    plt.xlabel('Time(s)')
    plt.ylabel('D_vel(m/s)')
    ax.yaxis.set_major_locator(y_major_locator)
    plt.ylim(0,1.5)
    plt.figure(3)
    ax.yaxis.set_major_locator(y_major_locator)
    plt.ylim(0,1.5)
    plt.plot(x1,Sys_Rate_Model_Helth_group)
    plt.xlabel('Time(s)')
    plt.ylabel('Sys_Health')
    plt.show()